import openpyxl
from openpyxl.styles import PatternFill

#To get row count
def getRowCount(file, sheetName):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[sheetName]
    return (sheet.max_row)

#To get column count
def getColumnCount(file, sheetName):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[sheetName]
    return (sheet.max_Column)

#To read data from excel
def readData(file, sheetName, rownum, columnnum):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[sheetName]
    return sheet.cell(rownum,columnnum).value

#To write data to excel
def writeData(file, sheetName, rownum, columnnum, data):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[sheetName]
    sheet.cell(rownum,columnnum).value = data
    workbook.save(file)

#To fill green colour for pass case
def fillGreenColour(file, sheetName, rownum, columnnum):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    greenfill= PatternFill(start_color='60b212',end_color='60b212',fill_type='solid')
    sheet.cell(rownum,columnnum).fill=greenfill
    workbook.save(file)

#To fill red colour for failed case
def fillRedColour(file, sheetName, rownum, columnnum):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    redfill = PatternFill(start_color='ff0000', end_color='ff0000', fill_type='solid')
    sheet.cell(rownum, columnnum).fill = redfill
    workbook.save(file)
