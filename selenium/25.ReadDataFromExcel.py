import openpyxl
#openpyxl - to import data from excel.xlsx
#install openpyxl in python interpreter

file="D:\\TESTING\\PythonAutomationProjects\\Testing.xlsx"
workbook=openpyxl.load_workbook(file)
sheet=workbook["Test1"]
rows=sheet.max_row #count no of rows
columns=sheet.max_column #count no of columns

#Read all rows and columns from Excel
for r in range(1,rows+1):
    for c in range(1, columns+1):
        print(sheet.cell(r,c).value, end='    ')
    print()
