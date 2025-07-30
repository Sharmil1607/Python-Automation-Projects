#Same Data
import openpyxl

#Single Data in multiple Rows and Columns
file="D:\\TESTING\\PythonAutomationProjects\\Testing.xlsx"
workbook=openpyxl.load_workbook(file)
sheet=workbook["Test2"]

for r in range(1,6):
    for c in range(1,4):
        sheet.cell(r,c).value="Welcome"
workbook.save(file)
print("1")

#Multiple Datas
file="D:\\TESTING\\PythonAutomationProjects\\Testing.xlsx"
workbook=openpyxl.load_workbook(file)
sheet=workbook["Test3"]

sheet.cell(1,1).value="PlayerName"
sheet.cell(1,2).value="Position"
sheet.cell(1,3).value="JerseyNumber"

sheet.cell(2,1).value="Kholi"
sheet.cell(2,2).value="3rdDown"
sheet.cell(2,3).value=18

sheet.cell(3,1).value="Sachin"
sheet.cell(3,2).value="Opener"
sheet.cell(3,3).value=10

sheet.cell(4,1).value="Dhoni"
sheet.cell(4,2).value="WicketKeeper"
sheet.cell(4,3).value=7

workbook.save(file)
print("2")